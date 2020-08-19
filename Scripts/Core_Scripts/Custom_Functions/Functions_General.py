import Core_Scripts.Custom_Functions.Functions_Lists as listFuncs
import pandas as pd
import os
import datetime


# get list of files from directory provided
def get_files_from_directory(directory):
    filelist = []
    # the os.walk method had three outputs shown below for each file and item it finds within the provided path
    for (directorypath, directorynames, filenames) in os.walk(directory):
        filelist.append(filenames)
        # can add a break if you dont want it to go into subfolders past the current one (you do, otherwise open fails)
        break
    # flatten f because it comes out as a list of lists of lists
    filelist = listFuncs.list_flatten(filelist)
    return filelist


def check_string_for_match(strval, strlist, caps_sensitive=True):
    # can also use something like strval in '\z'.join(strlist)
    logiclist = []
    if not caps_sensitive:
        for check in strlist:
            logiclist.append(str(strval).lower() == str(check).lower())
    else:
        for check in strlist:
            logiclist.append(strval == check)
    return any(logiclist)


def set_row_nones(list_for_length, string_to_set):
    # list only needed to calculate the length of how many rows to add
    stringmtx = []
    for val in list_for_length:
        stringmtx.append(string_to_set)
    return stringmtx


def max_of_objects(userlist, userobject=float):
    # only checks the max of the object type provided and requires a list to run
    # initiate is set to an unrealistic date to
    if userobject is float:
        initiate = 0
    elif userobject is datetime.datetime:
        initiate = datetime.datetime(1, 1, 1, 1, 1, 1)
    else:
        initiate = str(0)
    # if userobject is datetime.datetime:
    #     initiate = datetime.datetime(1, 1, 1, 00, 00, 00, 00)
    maximum = initiate
    for eachitem in userlist:
        if type(eachitem) is userobject:
            try:
                if maximum <= eachitem:
                    maximum = eachitem
            except TypeError:
                pass
    if maximum == initiate:
        return None
    else:
        return maximum


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    'from stackoverflow https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas'
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    # try:
    #     FileNotFoundError
    # except NameError:
    #     FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()